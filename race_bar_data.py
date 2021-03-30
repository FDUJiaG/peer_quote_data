import os
import string
import time
from datetime import datetime
import numpy as np
import pandas as pd
from common_utils import col_temp, state_dict, kcb_col, cyb_col, race_member_list
from op_ns_data import print_info, find_file_path, get_col_list, get_ns_info_data, judge_df, get_df_group
from op_ns_data import get_note, save_df
from op_ns_data import op_ns_data
import openpyxl
from openpyxl.styles import Font, PatternFill
import warnings

warnings.filterwarnings("ignore")


def get_race_bar_data():
    root_path = os.path.abspath(".")
    data_dir = os.path.join(root_path, "raw_data")
    output_dir = os.path.join(root_path, "output")
    save_name = "注册制同行报价汇总表"
    # file_name = ["科创板", "创业板"]
    file_name = ["创业板"]
    file_type = ".xlsx"
    data_name = "同行报价"
    sheet_name = "基础数据"
    my_comp_name = "上海迎水投资管理有限公司"

    df_col = [
        "股票名称",
        "询价日",
        "投资者名称",
        "申购价格",
        "备注"
    ]

    df_zero = pd.DataFrame(columns=df_col)

    for file in file_name:
        file_path = find_file_path(data_dir, file, file_type)
        if not file_path:
            return False

        raw_df = get_ns_info_data(file_path, sheet_name)
        if type(raw_df) is bool:
            return raw_df

        if file == file_name[-1]:
            raw_df = raw_df[cyb_col]
        elif file == file_name[0]:
            raw_df = raw_df[kcb_col]

        raw_df.columns = df_col
        df_zero = pd.concat([df_zero, raw_df])

    df_zero.sort_values(by=["询价日", "股票名称"], inplace=True, ascending=[False, True])
    print(df_zero.head())

    tzz_mc, sg_jg = df_col[2], df_col[3]

    # 获取全部的股票列表
    stock_list = df_zero[df_col[0]].tolist()
    union_stock_col = list(set(stock_list))
    union_stock_col.sort(key=stock_list.index)
    print(print_info(), end=" ")
    print("Get the stock list: \n{}".format(union_stock_col))

    # 获取无重复全部的投资者
    all_tzz_col = df_zero[tzz_mc].tolist()
    union_tzz_col = list(set(all_tzz_col))
    tzz_col_len = len(union_tzz_col)
    print(print_info(), end=" ")
    print("Get the tzz list: \n{}".format(union_tzz_col))

    race_member_list_copy = race_member_list.copy()
    print(race_member_list_copy)

    for race_item in race_member_list:
        print(race_item + "基金")
        count = 0
        attribute = 0
        for tzz_item in union_tzz_col:
            if set(race_item + "基金").issubset(set(tzz_item)):
                count += 1
                attribute = 1
            # elif set(race_item).issubset(set(tzz_item)):
            #     attribute = 2
            #     count += 1
                print(race_item, count, attribute)
                # race_member_list_copy.remove(race_item)
    print(race_member_list_copy)

    df_dict = dict()
    price_dict = dict()
    font_dict = dict()

    # for stock in union_stock_col:
    #     print(print_info(), end=" ")
    #     print("Operator the stock: {}".format(stock))
    #     df_dict[stock] = df_zero[df_zero[df_col[0]] == stock]
    #     df_group = get_df_group(df_dict[stock], tzz_mc, sg_jg)
    #     if type(df_group) is bool:
    #         return df_group
    #
    #     price_line = list()
    #     note_line = list()
    #     for tzz_item in union_tzz_col:
    #         if tzz_item in df_group.index:
    #             price_line.append(df_group[df_col[3]][tzz_item])
    #             note_line.append(get_note(df_dict[stock], tzz_mc, tzz_item, tzz_item, state_dict)[0])
    #         else:
    #             price_line.append("")
    #             note_line.append("")
    #
    #     price_dict[stock] = price_line
    #     font_dict[stock] = note_line
    #
    # # 对于单支股票，输出一行报价记录，以及一个用于标记颜色的单条记录
    # print(print_info(), end=" ")
    # print("Get the price dict:\n{}".format(price_dict))
    # print(print_info(), end=" ")
    # print("Get the font dict:\n{}".format(font_dict))
    #
    # TF = save_all_data(output_dir, save_name, file_type, price_dict, font_dict, union_tzz_col, tzz_col_len)

    return TF

    # try:
    #     save_df(df_note, root_path, file_path, sheet_name, op_file_type=".xlsx")
    #     return True
    # except:
    #     return False


def save_all_data(o_path, s_name, op_file_type, p_dict, f_dict, u_tzz_col, tzz_len):
    temp_df = pd.DataFrame.from_dict(p_dict)
    s_df = pd.DataFrame(temp_df.values.T, index=temp_df.columns, columns=u_tzz_col)
    op_date = time.strftime('%Y%m%d', time.localtime(time.time()))
    file_name = s_name + "_" + op_date + op_file_type
    save_path = os.path.join(o_path, file_name)
    try:
        writer = pd.ExcelWriter(save_path)
        if os.path.exists(writer.path):
            book = openpyxl.load_workbook(writer.path)
            writer.book = book
            # print(book.sheetnames)
        # if sheet_n in writer.book.sheetnames:
        #     writer.remove(writer[sheet_n])
        s_df.to_excel(writer)
        writer.save()
        writer.close()
        print(print_info(), end=" ")
        print("Save to the path: {}".format(save_path))
    except:
        print(print_info("E"), end=" ")
        print("Can not save to the path: {}".format(save_path))
        return False

    font_judge = set_font(f_dict, tzz_len, save_path)
    if not font_judge:
        return False

    return True


def set_font(f_dict, tzz_len, excel_name):
    # try:
    book = openpyxl.load_workbook(excel_name)
    sheet = book["Sheet1"]
    f_keys = list(f_dict.keys())
    f_keys_len = len(f_keys)

    for idx in range(1000):
        sheet.cell(1, idx + 1).fill = PatternFill(fill_type='solid', fgColor="F4A460")

    for idx, stock in zip(range(f_keys_len), f_keys):
        # 对于每一支股票逐行染色
        f_stock = f_dict[stock]
        for jdx, tzz_item in zip(range(tzz_len), f_stock):
            cell = sheet.cell(idx + 2, jdx + 2)
            if tzz_item == state_dict["低"]:
                cell.font = Font(color="008000")
                cell.fill = PatternFill(fill_type='solid', fgColor="00FA9A")
            elif tzz_item == state_dict["高"]:
                cell.font = Font(color="DC143C")
                cell.fill = PatternFill(fill_type='solid', fgColor="FFC0CB")
            elif tzz_item == state_dict["无"]:
                cell.font = Font(color="696969")
                cell.fill = PatternFill(fill_type='solid', fgColor="C0C0C0")

    book.save(excel_name)
    book.close()
    print(print_info(), end=" ")
    print("Font set!")
    # except:
    #     print(print_info("E"), end=" ")
    #     print("Can not set the font")
    #     return False
    return True


if __name__ == '__main__':
    TF = get_race_bar_data()
    if TF:
        print(print_info("S"), end=" ")
        print("Success!")
    else:
        print(print_info("E"))
        print("Error!")
