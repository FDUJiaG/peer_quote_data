import os
import string
import time
from datetime import datetime
import numpy as np
import pandas as pd
from common_utils import col_temp, state_dict, attention_dict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from WindPy import w
import warnings

warnings.filterwarnings("ignore")


def op_ns_data(file_name):
    root_path = os.path.abspath(".")
    data_dir = os.path.join(root_path, "raw_data")
    file_type = ".xlsx"
    data_name = "同行报价"
    sheet_name = "关注"
    col_list = get_col_list(root_path, data_name, sheet_name, file_type)
    if not col_list:
        return False

    file_path = find_file_path(data_dir, file_name, file_type)
    if not file_path:
        return False

    long_name = os.path.split(file_path)[-1].split(".")[0]
    ipo_name, ipo_code = get_name_and_code(long_name)
    # if ipo_code != "":
    #     data = w.wsd(
    #         ipo_code, "sec_name,ipo_inq_enddate",
    #         "ED-1TD", datetime.now().strftime("%Y-%m-%d")
    #     )
    #     if ipo_name != data.Data[0][0].split("-")[0]:
    #         if
    #         print(print_new_info("E", "R"), end=" ")
    #         print("Name: {} and Code: {} not match!".format(ipo_name, ipo_code))
    #         return False

    raw_df = get_ns_info_data(file_path)
    if type(raw_df) is bool:
        return raw_df

    # 替换
    raw_df.replace("招商基金管理有限公司", "招商基金", inplace=True)

    if not judge_df(raw_df):
        return False

    raw_df_key = raw_df.keys().tolist()
    for item in raw_df_key:
        if "价格" in item:
            sg_jg = item
        elif "投资者" in item or "交易员" in item:
            tzz_mc = item

    df_group = get_df_group_all(raw_df, tzz_mc, sg_jg)

    if type(df_group) is bool:
        return df_group

    tzz_list = list(set(df_group[tzz_mc].tolist()))
    df_note = output_df(tzz_mc, sg_jg, raw_df, df_group, tzz_list, col_list, col_temp, ipo_name, ipo_code)

    try:
        save_df(df_note, root_path, file_path, sheet_name, op_file_type=".xlsx")
        return True
    except:
        return False


def get_time(date=False, utc=False, msl=3):
    if date:
        time_fmt = "%Y-%m-%d %H:%M:%S.%f"
    else:
        time_fmt = "%H:%M:%S.%f"

    if utc:
        return datetime.utcnow().strftime(time_fmt)[:(msl - 6)]
    else:
        return datetime.now().strftime(time_fmt)[:(msl - 6)]


def print_info(status="I"):
    return "\033[0;33;1m[{} {}]\033[0m".format(status, get_time())


def print_new_info(status="I", color="Y"):
    if color == "R":
        return "\033[0;31;1m[{} {}]\033[0m".format(status, get_time())
    else:
        return "\033[0;33;1m[{} {}]\033[0m".format(status, get_time())


def six_code_to_ts_code(code):
    # 将股票代码转换成标准格式
    code = str(code)
    code = code.split(".")[0]

    if len(code) == 6 and code.isdigit():
        if code[0] == "6":
            exc_label = "SH"
        elif code[0] in ["0", "3"]:
            exc_label = "SZ"
        else:
            print(print_info("W"), end=" ")
            print("Can not get the Exchange Info from the code: {}".format(code))
            return False
        return ".".join([code, exc_label])
    else:
        print(print_new_info("E", "R"), end=" ")
        print("The stock code: {} is Error! Please check again!".format(code))
        return False


def get_name_and_code(long_n):
    # 拆分文件名成股票名称和代码
    ipo_c = ""
    if "_" in long_n:
        ipo_n = long_n.split("_")[0]
        ipo_c = long_n.split("_")[-1].split("(")[0]
        ipo_c = six_code_to_ts_code(ipo_c)
    else:
        ipo_n = long_n.split("(")[0]
    return ipo_n, ipo_c


def find_file_path(data_path, file_n, f_type=".xlsx"):
    if not os.path.exists(data_path):
        print(print_new_info("E", "R"), end=" ")
        print("The data directory: {} is not existed, please check again!".format(data_path))
        return False
    else:
        print(print_info(), end=" ")
        print("Find the data in path: {}".format(data_path))

        # 仅考虑符合条件格式的文件列表
        dir_list = os.listdir(data_path)
        f_type_temp = f_type.split(".")[-1]
        dir_list = [item for item in dir_list if item.split(".")[-1] == f_type_temp]

        file_path = list()

        for item in dir_list:
            if file_n in item and "$" not in item:
                file_path.append(item)

        if len(file_path) == 0:
            print(print_new_info("E", "R"), end=" ")
            print("No {} file in the data path: {}".format(file_n, data_path))
            return False
        elif len(file_path) > 2:
            print(print_new_info("E", "R"), end=" ")
            print("Not only one file include {} in the data path: {}".format(file_n, data_path))
            return False
        else:
            file_path = os.path.join(data_path, file_path[0])
            print(print_info(), end=" ")
            print("Get the file path: {}".format(file_path))

    return file_path


def get_col_list(r_path, data_n, sheet_n, f_type=".xlsx"):
    file_list = os.listdir(r_path)
    data_list = [
        file for file in file_list if data_n in file and f_type.split(".")[-1] in file.split(".")[-1]
        and "~$" not in file
    ]

    print(print_info(), end=" ")
    print("Data List: {}".format(data_list))

    if len(data_list) > 0:
        data_path = data_list[-1]
        try:
            col_list = pd.read_excel(os.path.join(r_path, data_path), sheet_name=sheet_n, header=0).columns.tolist()
            print(print_info(), end=" ")
            print("Columns get!")
            print(col_list)
        except:
            print(print_new_info("E", "R"), end=" ")
            print("Can not get the columns!")
            return False
    else:
        print(print_info(), end=" ")
        print("No column data found!")
        return False

    return col_list


def get_ns_info_data(f_path, sheet_n="Sheet1"):
    try:
        df = pd.read_excel(f_path, sheet_name=sheet_n, header=0)
        print(print_info(), end=" ")
        print("Successfully loading the file: {}".format(f_path))
        return df
    except:
        print(print_new_info("E", "R"), end=" ")
        print("Could not load the file: {}".format(f_path))
        return False


def judge_df(df):
    judge_arr = np.where(np.diff(df["序号"].tolist()) != 1)

    if len(judge_arr[0]) > 0:
        print(print_info("W"), end=" ")
        print("Problems with documentation items, please check again! The warning item idx: {}.".format(
            (judge_arr[0] + 1).tolist()
        ))
        return False
    else:
        return True


def get_df_group_all(df, index_n, price):
    # 全部报价信息，并非只显示报价众数
    # 投资者名称，报价，备注，计数
    try:
        index_name = index_n
        df_group_all = df.groupby([index_name, price, "备注"])["序号"].count().reset_index()
        df_group_all.rename(columns={"序号": "计数"}, inplace=True)
        df_group_all.sort_values(by=[index_name, "计数"], ascending=[True, False], inplace=True)
        print(print_info(), end=" ")
        print("The Group by DataFrame:\n {}".format(df_group_all))
        # df_group_all.to_excel("test.xlsx")
        return df_group_all
    except:
        print(print_new_info("E", "R"), end=" ")
        print("There is something wrong in DataFrame:\n {}".format(df.head()))
        return False


def get_note(df, tzz_mc, short_title, tzz_n, s_dict):
    df_temp = df[df[tzz_mc] == tzz_n]
    note_dict = dict(df_temp["备注"].value_counts())
    note_key = list(note_dict.keys())
    note_sum = sum(note_dict.values())
    s_key = list(s_dict.keys())

    # 默认“有效报价”不标记
    note = ""
    desc_note = ""
    valid_len = 0
    for item in note_key:
        if s_key[-1] in item or "入围" == item:
            valid_len = note_dict[item]
            note_key.remove(item)

    note_len = len(note_key)

    if note_len:
        # 清洗一个备注情况的字典，统计低价，高价，无效的情况
        value_dict = dict()
        for idx in range(len(s_key) - 1):
            for item in note_key:
                if s_key[idx] in item:
                    value_dict[s_key[idx]] = 0
        for item in note_key:
            for idx in range(len(s_key) - 1):
                if s_key[idx] in item:
                    value_dict[s_key[idx]] += note_dict[item]

        # 单个投资者的报价标记，注意优先级
        value_key = list(value_dict.keys())
        if s_key[0] in value_key:
            note = s_dict[s_key[0]]
        elif s_key[1] in value_key:
            note = s_dict[s_key[1]]

        # 其他种种失败的报价情形
        elif s_key[2] in value_key or len(value_key) == 0:
            note = s_dict[s_key[2]]
        else:
            print(print_info("W"), end=" ")
            print("Unexpected events occurred in {}!".format(tzz_n))

        # 仅有一种非有效报价情况
        if len(value_key) == 1:
            if valid_len == 0:
                desc_note += short_title + str(note_sum) + "个" + s_dict[value_key[0]]
            else:
                desc_note += short_title + str(note_sum) + "个配售对象中" \
                             + str(value_dict[value_key[0]]) + "个" + s_dict[value_key[0]]

        # 两种以上非有效报价情况
        else:
            desc_note += short_title + str(note_sum) + "个配售对象中"
            for v_item in value_key:
                desc_note += str(value_dict[v_item]) + "个" + s_dict[v_item]

    return note, desc_note


def easy_check(staus):
    if "高" in staus:
        return state_dict["高"]
    elif "低" in staus:
        return state_dict["低"]
    elif "未" in staus or "无" in staus:
        return state_dict["无"]
    else:
        return state_dict["有"]


def output_df(tzz_mc, sg_jg, raw_df, df_group, tzz_list, col_list, col_temp, ipo_n, ipo_c):
    # 1有效, 2低剔, 3高剔
    df_output = pd.DataFrame(columns=col_temp)
    desc_list = list()

    op_list = col_list[:6] + list(attention_dict.keys())
    op_list.append("备注")

    df_note = pd.DataFrame(columns=col_list)

    for item in op_list:
        p = 0
        if item in attention_dict.keys():
            tzz_item = attention_dict[item]
            if tzz_item in tzz_list:
                p = df_group[df_group[tzz_mc] == tzz_item][sg_jg].tolist()
                info = df_group[df_group[tzz_mc] == tzz_item]["备注"].tolist()
                info = [easy_check(info_item) for info_item in info]
                # print(tzz_item, info, p)
                p_zip = zip(p, info)

                right_p = list()
                low_p = list()
                high_p = list()
                for p_item, i_item in p_zip:
                    if i_item == "高价剔除":
                        high_p.append(str(p_item))
                    elif i_item == "低价剔除":
                        low_p.append(str(p_item))
                    elif i_item == "有效报价":
                        right_p.append(str(p_item))
                h_p = "\n".join(high_p)
                l_p = "\n".join(low_p)
                r_p = "\n".join(right_p)

                if r_p != "":
                    df_note[item + "1"] = [r_p, ""]
                if l_p != "":
                    df_note[item + "2"] = [l_p, "低价剔除"]
                if h_p != "":
                    df_note[item + "3"] = [h_p, "高价剔除"]

                item_temp, desc_item = get_note(raw_df, tzz_mc, item, tzz_item, state_dict)
        if p == 0:
            if item == "证券名称":
                df_note[op_list[0]] = [ipo_n, ipo_n]
            elif item == "证券代码":
                # output_item[col_temp[2]] = ipo_c
                df_note.loc[0, op_list[1]] = ipo_c
            elif item == "询价日期":
                if ipo_c != "":
                    data = w.wsd(
                        ipo_c, "sec_name,ipo_inq_enddate",
                        "ED-1TD", datetime.now().strftime("%Y-%m-%d")
                    )
                    xj_date = data.Data[-1][0].strftime("%Y-%m-%d")
                    df_note.loc[0, op_list[2]] = xj_date
            elif item == "我司报价":
                fullname = "上海迎水投资管理有限公司"
                try:
                    p = df_group[df_group[tzz_mc] == fullname][sg_jg].tolist()[0]
                except:
                    p = ""
                df_note.loc[0, "我司报价"] = p
                item_temp, desc_item = get_note(raw_df, tzz_mc, item, fullname, state_dict)
                df_note.loc[1, "我司报价"] = item_temp
        else:
            if desc_item != "":
                desc_list.append(desc_item)

    desc_list_unique = list(set(desc_list))
    desc_list_unique.sort(key=desc_list.index)
    desc_note_unique = "；".join(desc_list_unique)
    print(print_info(), end=" ")
    print("Note: {}".format(desc_note_unique))
    df_note.loc[0, "备注"] = desc_note_unique
    # df_note.to_excel("df_note.xlsx")
    return df_note


def save_df(df, r_path, f_path, sheet_n, op_file_type=".xlsx"):
    op_date = time.strftime('%Y%m%d', time.localtime(time.time()))
    zq_name = os.path.split(f_path)[-1].split(".")[0]
    if len(zq_name) <= 4:
        file_name = zq_name + "_" + op_date + op_file_type
    else:
        file_name = zq_name[:4] + "_" + op_date + op_file_type
    # 原逻辑是建立两张表，已优化更新
    # if "全部" in sheet_n:
    #     file_name = os.path.split(f_path)[-1].split(".")[0] + "_all_" + op_date + op_file_type
    # else:
    #     file_name = os.path.split(f_path)[-1].split(".")[0] + "_" + op_date + op_file_type
    output_dir = os.path.join(r_path, "output")
    if not os.path.exists(output_dir):
        os.makedirs(os.path.join(r_path, "output"))
        print(print_info(), end=" ")
        print("Created the dir: {}".format(output_dir))
    save_path = os.path.join(output_dir, file_name)
    try:
        writer = pd.ExcelWriter(save_path)
        if os.path.exists(writer.path):
            book = openpyxl.load_workbook(writer.path)
            writer.book = book
            # print(book.sheetnames)
        if sheet_n in writer.book.sheetnames:
            writer.remove(writer[sheet_n])
        df[:-1].to_excel(writer, sheet_n, index=None)
        writer.save()
        writer.close()
        print(print_info(), end=" ")
        print("Save to the path: {}".format(save_path))
    except:
        print(print_new_info("E", "R"), end=" ")
        print("Can not save to the path: {}".format(save_path))

    font_judge = set_font(df, save_path, sheet_n)
    if not font_judge:
        return False
    return True


def set_font(df, excel_name, sheet_name):
    try:
        book = openpyxl.load_workbook(excel_name)
        sheet = book[sheet_name]
        shape = df.shape
        label_list = df.iloc[-1].tolist()

        # 备注那列宽一点
        if sheet_name == "关注":
            sheet.column_dimensions[get_column_letter(shape[1])].width = 108
            wrap = True
        else:
            wrap = False

        for idx, label in zip(range(shape[1]), label_list):
            cell = sheet.cell(shape[0], idx + 1)
            cell.alignment = Alignment(wrapText=wrap, vertical="center")
            if idx >= 6 and idx % 3 == 0 and sheet_name == "关注":
                cell.border = Border(left=Side(border_style='medium', color='000000'))
            if label == state_dict["低"]:
                cell.font = Font(color="008000")
                cell.fill = PatternFill(fill_type='solid', fgColor="00FA9A")
            elif label == state_dict["高"]:
                cell.font = Font(color="DC143C")
                cell.fill = PatternFill(fill_type='solid', fgColor="FFC0CB")
            # 自2021年3月23日起，无需标注无效报价
            # elif label == state_dict["无"]:
            #     cell.font = Font(color="696969")
            #     cell.fill = PatternFill(fill_type='solid', fgColor="C0C0C0")
        book.save(excel_name)
        book.close()
        print(print_info(), end=" ")
        print("Font set!")
        return True
    except:
        print(print_new_info("E", "R"), end=" ")
        print("Can not set the font")
        return False


if __name__ == '__main__':
    w.start()
    w.isconnected()
    TF = op_ns_data("奥尼电子_301189.xlsx")
    if TF:
        print(print_info("S"), end=" ")
        print("Success!")
    else:
        print(print_new_info("E", "R"), end=" ")
        print("Error!")
    w.close()
