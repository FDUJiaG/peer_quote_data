import os
import json
import string
import time
from datetime import datetime
import numpy as np
import pandas as pd
from common_utils import color_base
from op_ns_data import print_info, print_new_info, get_ns_info_data
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import warnings

warnings.filterwarnings("ignore")


def op_top_comp(d_dir, t_rate, f_type):
    sub_dir = "history"
    drop_name = "基础数据"
    file_list = get_file_list(d_dir, sub_dir, f_type, drop_name)
    print(print_new_info(), end=" ")
    print("Get the file list:\n {}".format(file_list))

    base_name = "同行报价"
    sheet_name = "全部"
    base_col = ["证券名称", "证券代码", "询价日期", "发行价", "高剔价格"]
    base_data = get_base_data(root_path, base_name, sheet_name, base_col, file_type)
    if type(base_data) is bool and not base_data:
        return False

    # 构建数据框
    state_left = base_col.copy()
    state_left.insert(3, "2020EPS")
    state_left.append("备注")
    df_data = pd.DataFrame(data=state_left[1:], columns=["证券名称"])
    base_len = df_data.shape[0] + 1
    color_dict = dict()

    eps_name = "注册制发行价预测"
    eps_sheet_name = "Sheet1"
    eps_col = ["证券名称", "证券代码", "2020EPS", "2021EPS"]
    eps_data = get_base_data(root_path, eps_name, eps_sheet_name, eps_col, file_type, 1)
    if type(eps_data) is bool and not eps_data:
        return False
    # 去除重复项
    eps_data.drop_duplicates(subset="证券代码", inplace=True)

    # 获取处理字典路径 file:path
    op_stock_dict = dict()
    for file_item in file_list:
        # 获取标的名称
        k_tmp = file_item.split("\\")[-1].split(".")[0].split("_")[0]
        op_stock_dict[k_tmp] = file_item

    # 根据同行报价的股票顺序依次检索
    for stock_item in base_data.index:
        if stock_item not in op_stock_dict:
            # 如果找不到原始数据，暂时不处理
            # print(stock_item)
            continue

        # 发行价, 高剔价格
        ipo_price = base_data["发行价"].loc[stock_item]
        high_drop_price = base_data["高剔价格"].loc[stock_item]

        # 处理数据
        file_path = os.path.join(d_dir, op_stock_dict[stock_item])
        df_raw = get_ns_info_data(file_path)
        print(print_new_info(), end=" ")
        print("Get the raw dataframe:\n{}".format(df_raw.head()))

        # 计算 eps
        eps = eps_data["2020EPS"].loc[stock_item]

        # 获取分类数据
        df_group = get_sort_df(df_raw)
        sum_state, count_state, sum_color, count_color = get_state(df_group, t_rate, ipo_price, eps)

        head_list = [
            base_data["证券代码"].loc[stock_item],   # 证券代码
            base_data["询价日期"].loc[stock_item],   # 询价日期
            eps,
            "{:.2f}({}pe:{:.2f})".format(ipo_price, 20, ipo_price / eps),                  # 发行价
            "{:.2f}({}pe:{:.2f})".format(high_drop_price, 20, high_drop_price / eps),      # 高剔价格
            sum_state      # 备注信息
        ]
        # axis=1 表示行对齐
        df_data = pd.concat(
            [
                df_data,
                pd.DataFrame(
                    data=head_list + list(sum_color.keys()), columns=[stock_item])
            ],
            axis=1
        )
        color_dict[stock_item] = list(sum_color.values())

    op_date = time.strftime('%Y%m%d', time.localtime(time.time()))
    output_name = "同行报价投资者Top榜单_{}.{}".format(op_date, f_type)
    output_path = os.path.join(os.path.abspath("."), "output", output_name)
    df_data.to_excel(output_path, index=None)
    print(print_new_info(), end=" ")
    print("Successfully Saved to: {}".format(output_path))
    set_color(color_dict, output_path, base_len)

    return True


def get_base_data(r_path, data_n, sheet_n, col_list, f_type="xlsx", h_no=0):
    # 获取同行报价数据，输出代码，询价日期，发行价，高剔价格等信息
    file_list = os.listdir(r_path)
    data_list = [
        file for file in file_list if data_n in file and "~$" not in file and f_type in file.split(".")[-1]
    ]

    print(print_info(), end=" ")
    print("Data List: {}".format(data_list))

    if len(data_list) > 0:
        data_path = data_list[-1]
        try:
            base_data = pd.read_excel(os.path.join(r_path, data_path), sheet_name=sheet_n, header=h_no)
            base_data = base_data[col_list]
            base_data.set_index("证券名称", inplace=True)
            print(print_info(), end=" ")
            print("Base Data Get!")
            print(base_data)
        except:
            print(print_new_info("E", "R"), end=" ")
            print("Can not get the base data!")
            return False
    else:
        print(print_info(), end=" ")
        print("No column data found!")
        return False

    return base_data


def get_file_list(d_dir, s_dir, f_type, dp_name):
    # 获取文件夹下所有需要处理的文件
    dir_list = os.listdir(d_dir)
    file_list = dir_list.copy()
    out_list = list()

    if type(s_dir) is str and s_dir in dir_list:
        file_list.remove(s_dir)
        sub_path = os.path.join(d_dir, s_dir)
        sub_list = os.listdir(sub_path)
        sub_list = [os.path.join(s_dir, item) for item in sub_list]
        file_list.extend(sub_list)
    elif type(s_dir) is list:
        for s_item in dir_list:
            if s_item in file_list:
                file_list.remove(s_dir)
                sub_path = os.path.join(d_dir, s_dir)
                sub_list = os.listdir(sub_path)
                sub_list = [os.path.join(s_item, item) for item in sub_list]
                file_list.extend(sub_list)

    for f_item in file_list:
        if f_item.split(".")[-1] == f_type and "~$" not in f_item and dp_name not in f_item:
            out_list.append(f_item)

    return out_list


def get_std_col(r_df):
    # 获取一个标准的数据框列名，主要获取投资者名称，申购价格，申购股数
    col_list = list()
    tmp_list = r_df.columns
    for tmp in tmp_list:
        if "名称" in tmp and "对象" not in tmp:
            col_list.append(tmp)
        elif "价格" in tmp:
            col_list.append(tmp)
        elif "数量" in tmp:
            col_list.append(tmp)
    return col_list


def get_sort_df(r_df):
    # 处理数据框, 以名称和价格的元组对为key, 聚合报价的总申购股数和产品数量之和
    col_list = get_std_col(r_df)
    # print(col_list)
    df_group = r_df.groupby([col_list[0], col_list[1]])[col_list[2]].agg([np.sum, np.count_nonzero]).reset_index()
    group_col = df_group.columns
    # 排序规则是价格降序优先，总申购股数降序其次，产品数量之和升序再次
    df_group.sort_values(by=[col_list[1], group_col[-2], group_col[-1]], ascending=[False, False, True], inplace=True)
    df_group["sum_pct"] = round(df_group[group_col[-2]] / sum(df_group[group_col[-2]]) * 10000, 2)
    df_group["count_pct"] = round(df_group[group_col[-1]] / sum(df_group[group_col[-1]]) * 10000, 2)
    df_group["sum_pct_cs"] = df_group["sum_pct"].cumsum()
    df_group["count_pct_cs"] = df_group["count_pct"].cumsum()
    return df_group


def get_state(g_df, t_rate, ipo_p, eps_data):
    # 获取备注信息
    rate_len = len(t_rate)
    sum_cs, count_cs = g_df.columns[-2:]

    # 存贮位置信息
    sum_dict, count_dict = dict(), dict()
    # 总述的中间变量
    sum_tmp, count_tmp = list(), list()
    # 颜色字典
    sum_color, count_color = dict(), dict()

    for idx, r in enumerate(t_rate):
        # 找出符合比例的第一个到达索引，并反查数据
        s_idx = list(g_df[sum_cs] > r * 10000).index(True)
        sum_dict[r] = [s_idx, g_df[sum_cs].tolist()[s_idx]]
        flag_price = g_df[g_df.columns[1]].iloc[s_idx]      # 锚点价格
        c_idx = list(g_df[count_cs] > r * 10000).index(True)
        count_dict[r] = [c_idx, g_df[count_cs].tolist()[c_idx]]

        sum_tmp.append("前{:.1%}的公司展示至{:.2f}bp，与发行价差值{:.2f}".format(
            r, sum_dict[r][-1], flag_price - ipo_p))
        count_tmp.append("前{:.1%}的公司展示至{:.2f}bp".format(r, sum_dict[r][-1]))

    sum_state = "按照申购股数占比" + "；".join(sum_tmp)
    count_state = "按照申购产品数占比" + "；".join(count_tmp)

    # print(sum_dict, count_dict)

    print(print_new_info(), end=" ")
    print(sum_state)
    print(print_new_info(), end=" ")
    print(count_state)

    sum_dk = list()
    for item in sum_dict.values():
        sum_dk.append(item[0])
    for sk_id in range(sum_dk[-1] + 1):
        color_ct = rate_len + 1
        for r_idx in range(rate_len):
            if sk_id <= sum_dk[r_idx]:
                color_ct -= 1

        # print(g_df[g_df.columns[1]].tolist()[sk_id], eps_data)
        color_key = "{} 报价：{}(20pe:{:.2f})，占比{:.2f}bp".format(
            g_df[g_df.columns[0]].tolist()[sk_id],
            g_df[g_df.columns[1]].tolist()[sk_id],
            g_df[g_df.columns[1]].tolist()[sk_id] / eps_data,
            g_df["sum_pct"].tolist()[sk_id]
        )

        sum_color[color_key] = color_ct

    # 查看颜色对应字典
    # print(json.dumps(sum_color, indent=4, ensure_ascii=False))

    return sum_state, count_state, sum_color, count_color


def set_color(c_dict, excel_name, b_len, sheet_name="Sheet1"):
    left, right, top, bottom = [Side(style="thin", color="000000")] * 4
    try:
        book = load_workbook(excel_name)
        sheet = book[sheet_name]

        c_len = len(list(c_dict.keys()))

        # 控制第一列颜色
        for col_idx in range(b_len):
            cell = sheet.cell(col_idx + 1, 1)
            cell.font = Font(color="102B6A", bold=True)
            cell.fill = PatternFill(fill_type='solid', fgColor="7BBFEA")
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for s_idx, s_item in zip(range(1, c_len + 1), c_dict.keys()):
            # 循环每只股票
            # 设置列宽
            sheet.column_dimensions[get_column_letter(s_idx + 1)].width = 30
            # 控制表头
            for col_idx in range(b_len):
                cell = sheet.cell(col_idx + 1, s_idx + 1)
                if col_idx == 0:
                    cell.font = Font(color="102B6A", bold=True)
                else:
                    cell.font = Font(color="102B6A")
                cell.fill = PatternFill(fill_type='solid', fgColor="7BBFEA")
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # 控制具体颜色
            r_len = len(c_dict[s_item])
            for r_idx, r_item in zip(range(1, r_len + 1), c_dict[s_item]):
                # 循环每行, 根据表头行数修改
                cell = sheet.cell(r_idx + b_len, s_idx + 1)
                cell.font = Font(color=color_base[r_item][0])
                cell.fill = PatternFill(fill_type='solid', fgColor=color_base[r_item][1])
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        book.save(excel_name)
        book.close()
        print(print_info(), end=" ")
        print("Font set!")
    except:
        print(print_new_info("E", "R"), end=" ")
        print("Can not set the font")
        return False
    return True


if __name__ == '__main__':
    root_path = os.path.abspath(".")
    # data_dir = os.path.join(root_path, "raw_data_test")     # 测试
    data_dir = os.path.join(root_path, "raw_data")        # 生产
    file_type = "xlsx"
    rate_list = [0.01, 0.02, 0.03]
    op_top_comp(data_dir, rate_list, file_type)
